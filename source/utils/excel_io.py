import os
import pandas as pd
from datetime import datetime
from source.config.settings import INPUT_DIR, OUTPUT_DIR
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def read_input_excel(filename: str) -> pd.DataFrame:
    filepath = os.path.join(INPUT_DIR, filename)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Cannot find file input: {filepath}")

    df = pd.read_excel(filepath)
    print(f"[ExcelIO] Read {len(df)} row from {filepath}")
    return df

def write_output_excel(df: pd.DataFrame, base_filename: str = "compare_result"):
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_filename}_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    df.to_excel(filepath, index=False)
    print(f"[ExcelIO] Print results in {filepath}")
    return filepath

def write_output_excel_table(df: pd.DataFrame, base_filename: str = "compare_result"):
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_filename}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    df.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    # --- Style setup ---
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    red_fill = PatternFill(start_color="FAE2DE", end_color="FAE2DE", fill_type="solid")
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # --- Apply header style ---
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border

    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = gray_fill if idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill

    # --- Apply border + conditional red fill ---
    headers = [cell.value for cell in ws[1]]
    field_col = headers.index("Field") + 1 if "Field" in headers else None
    status_col = headers.index("Status") + 1 if "Status" in headers else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border

        if field_col and status_col:
            field_val = row[field_col - 1].value
            status_val = str(row[status_col - 1].value).lower() if row[status_col - 1].value else ""

            if status_val != "match" and field_val != "Summary":
                for cell in row:
                    cell.fill = red_fill

            if field_val == "Summary":
                for cell in row:
                    cell.font = Font(bold=True)

    # --- Auto filter ---
    ws.auto_filter.ref = ws.dimensions

    # --- Auto column width ---
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = min(max_length + 2, 100)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(filepath)
    print(f"[ExcelIO] Print results in {filepath}")
    return filepath

